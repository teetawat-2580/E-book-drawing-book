import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

def create_excel_tool():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # Styling Definitions
    # -------------------------------------------------------------
    font_title = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
    font_section = Font(name='Segoe UI', size=12, bold=True, color='1E293B')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_toggle = Font(name='Segoe UI', size=11, bold=True, color='0F172A')
    font_bold = Font(name='Segoe UI', size=11, bold=True)
    font_regular = Font(name='Segoe UI', size=10)
    font_code = Font(name='Consolas', size=10, color='0F766E')

    fill_title = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Deep Blue
    fill_header = PatternFill(start_color='0284C7', end_color='0284C7', fill_type='solid') # Ocean Blue
    fill_toggle_yes = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid') # Soft Green
    fill_toggle_hdr = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid') # Soft Yellow
    fill_sql_hdr = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid') # Teal Blue
    fill_sql_bg = PatternFill(start_color='F0FDFA', end_color='F0FDFA', fill_type='solid') # Mint Tint
    fill_accent = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Data Validation Yes/No
    dv_yes_no = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)

    # -------------------------------------------------------------
    # 1. INSERT Generator Sheet (MULTI-ROW INSERT WITHOUT ID)
    # -------------------------------------------------------------
    ws_insert = wb.create_sheet(title="INSERT_Generator")
    ws_insert.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_insert.merge_cells("A1:N1")
    ws_insert["A1"] = "🚀 MULTI-ROW BATCH INSERT GENERATOR (ไม่ใส่ ID ใช้ Auto-Increment ลดภาระเซิร์ฟเวอร์)"
    ws_insert["A1"].font = font_title
    ws_insert["A1"].fill = fill_title
    ws_insert["A1"].alignment = align_center
    ws_insert.row_dimensions[1].height = 40

    # Instructions
    ws_insert["A2"] = "คำแนะนำ: กรอกข้อมูลลงในแถวด้านล่าง (ไม่ต้องกรอก ID) ระบบจะรวมคำสั่งเป็น Multi-Row INSERT ให้อัตโนมัติในกล่องสีเขียวเข้มด้านล่าง"
    ws_insert["A2"].font = font_section

    # Multi-Row Master SQL Box in Row 3
    ws_insert.merge_cells("A3:M3")
    ws_insert["A3"] = "⚡ MULTI-ROW BATCH INSERT STATEMENT (คัดลอกกล่องนี้ไปใช้รันได้ทันทีทั้งชุด):"
    ws_insert["A3"].font = font_bold
    ws_insert["A3"].fill = fill_toggle_hdr

    ws_insert.merge_cells("A4:N4")
    # Using _xlfn.TEXTJOIN to prevent #NAME? error in Excel
    multi_insert_formula = (
        '="INSERT INTO books (title, author_name, publisher, category, price, original_price, pages_count, file_type, badge, cover_image_url, file_path, sample_file_path, description)" & CHAR(10) & '
        '"VALUES " & CHAR(10) & _xlfn.TEXTJOIN("," & CHAR(10), TRUE, N6:N55) & ";"'
    )
    ws_insert["A4"] = multi_insert_formula
    ws_insert["A4"].font = font_code
    ws_insert["A4"].fill = fill_sql_bg
    ws_insert["A4"].alignment = align_top_left
    ws_insert["A4"].border = border_thin
    ws_insert.row_dimensions[4].height = 75

    # Data Table Headers in Row 5 (Without ID column)
    headers_insert = [
        "title (ชื่อหนังสือ)*", "author_name (ผู้แต่ง)", "publisher (สำนักพิมพ์)", 
        "category (หมวดหมู่)*", "price (ราคาขาย)*", "original_price (ราคาเต็ม)", 
        "pages_count (จำนวนหน้า)", "file_type (ประเภทไฟล์)", "badge (ป้ายติด)", 
        "cover_image_url (URL ปก)", "file_path (URL ไฟล์เต็ม)", "sample_file_path (URL ไฟล์ตัวอย่าง)", 
        "description (รายละเอียด)", "📦 ROW TUPLE VALUE (Multi-Row Component)"
    ]

    ws_insert.row_dimensions[5].height = 28
    for col_num, h_text in enumerate(headers_insert, 1):
        cell = ws_insert.cell(row=5, column=col_num, value=h_text)
        cell.font = font_header
        cell.fill = fill_sql_hdr if col_num == 14 else fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # Sample Books Data (No ID column)
    sample_books_noid = [
        ["สมุดระบายสีสัตว์น่ารัก เล่ม 1", "คลังสมอง", "คลังสมอง KLANGSAMONG", "สมุดระบายสีเด็ก", 49, 99, "30 หน้า", "PDF", "ขายดี", "https://firebasestorage.googleapis.com/.../cover1.jpg", "https://firebasestorage.googleapis.com/.../book1.pdf", "https://firebasestorage.googleapis.com/.../sample1.pdf", "แบบฝึกหัดระบายสีสำหรับเด็กปฐมวัย เสริมสร้างพัฒนาการ"],
        ["ชีทคณิตศาสตร์ คิดเลขเร็ว เล่ม 1", "คลังสมอง", "คลังสมอง KLANGSAMONG", "คณิตศาสตร์", 59, 129, "45 หน้า", "PDF", "แนะนำ", "https://firebasestorage.googleapis.com/.../cover2.jpg", "https://firebasestorage.googleapis.com/.../book2.pdf", "https://firebasestorage.googleapis.com/.../sample2.pdf", "แบบฝึกหัดคิดเลขเร็ว บวก ลบ เลข 2 หลัก สำหรับเด็กประถม"],
        ["แฟลชการ์ด คำศัพท์ภาษาอังกฤษ 100 คำ", "คลังสมอง", "คลังสมอง KLANGSAMONG", "แฟลชการ์ด", 39, 79, "50 การ์ด", "PDF", "มาใหม่", "", "", "", "แฟลชการ์ดคำศัพท์ภาษาอังกฤษพร้อมรูปภาพประกอบน่ารัก"]
    ]

    for r_idx in range(6, 56): # Rows 6 to 55
        ws_insert.row_dimensions[r_idx].height = 22
        if r_idx - 6 < len(sample_books_noid):
            row_data = sample_books_noid[r_idx - 6]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_insert.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_regular
                cell.border = border_thin
        else:
            for c_idx in range(1, 14):
                ws_insert.cell(row=r_idx, column=c_idx).border = border_thin

        # Value tuple formula for Row r (Column N / Col 14)
        tuple_formula = (
            f'=IF(ISBLANK(A{r_idx}), "", '
            f'"(\'" & SUBSTITUTE(A{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(B{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(C{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(D{r_idx}, "\'", "\'\'") & "\', " & '
            f'IF(ISBLANK(E{r_idx}), "0", E{r_idx}) & ", " & '
            f'IF(ISBLANK(F{r_idx}), "NULL", F{r_idx}) & ", \'" & '
            f'SUBSTITUTE(G{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(H{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(I{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(J{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(K{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(L{r_idx}, "\'", "\'\'") & "\', \'" & '
            f'SUBSTITUTE(M{r_idx}, "\'", "\'\'") & "\')")'
        )
        cell_tuple = ws_insert.cell(row=r_idx, column=14, value=tuple_formula)
        cell_tuple.font = font_code
        cell_tuple.fill = fill_sql_bg
        cell_tuple.border = border_thin

    # -------------------------------------------------------------
    # 2. UPDATE Generator Sheet (Dynamic Column Selective Update)
    # -------------------------------------------------------------
    ws_update = wb.create_sheet(title="UPDATE_Generator")
    ws_update.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_update.merge_cells("A1:O1")
    ws_update["A1"] = "🛠️ DYNAMIC BATCH UPDATE SQL GENERATOR (เลือกคอลัมน์ที่ต้องการอัปเดต)"
    ws_update["A1"].font = font_title
    ws_update["A1"].fill = fill_title
    ws_update["A1"].alignment = align_center
    ws_update.row_dimensions[1].height = 40

    # Instructions & Controls
    ws_update["A2"] = "เลือก YES/NO ในบรรทัดที่ 3 เพื่อเลือกว่าจะอัปเดตคอลัมน์ใดบ้าง | ระบบจะสร้างคำสั่ง UPDATE books SET ... WHERE id=X ให้อัตโนมัติ"
    ws_update["A2"].font = font_section

    # Row 3: Toggle Headers ("UPDATE THIS FIELD?")
    ws_update.row_dimensions[3].height = 25
    ws_update.cell(row=3, column=1, value="🔑 Key ID (FIXED)").font = font_bold
    ws_update.cell(row=3, column=1).alignment = align_center
    ws_update.cell(row=3, column=1).fill = fill_toggle_hdr

    ws_update.add_data_validation(dv_yes_no)

    toggles_default = [
        ("title", "YES"), ("author_name", "NO"), ("publisher", "NO"), 
        ("category", "YES"), ("price", "YES"), ("original_price", "YES"), 
        ("pages_count", "NO"), ("file_type", "NO"), ("badge", "NO"), 
        ("cover_image_url", "YES"), ("file_path", "YES"), ("sample_file_path", "NO"), 
        ("description", "NO")
    ]

    for idx, (col_name, default_state) in enumerate(toggles_default, 2):
        cell = ws_update.cell(row=3, column=idx, value=default_state)
        cell.font = font_toggle
        cell.fill = fill_toggle_yes if default_state == "YES" else fill_accent
        cell.alignment = align_center
        cell.border = border_thin
        dv_yes_no.add(cell)

    ws_update.cell(row=3, column=15, value="⚙️ DYNAMIC UPDATE STATEMENT").font = font_bold
    ws_update.cell(row=3, column=15).alignment = align_center
    ws_update.cell(row=3, column=15).fill = fill_toggle_hdr

    # Row 4: Column Headers
    headers_update = [
        "id (ระบุ ID)*", "title (ชื่อหนังสือ)", "author_name (ผู้แต่ง)", "publisher (สำนักพิมพ์)", 
        "category (หมวดหมู่)", "price (ราคาขาย)", "original_price (ราคาเต็ม)", 
        "pages_count (จำนวนหน้า)", "file_type (ประเภทไฟล์)", "badge (ป้ายติด)", 
        "cover_image_url (URL ปก)", "file_path (URL ไฟล์เต็ม)", "sample_file_path (URL ตัวอย่าง)", 
        "description (รายละเอียด)", "⚡ GENERATED UPDATE SQL STATEMENT"
    ]

    ws_update.row_dimensions[4].height = 28
    for col_num, h_text in enumerate(headers_update, 1):
        cell = ws_update.cell(row=4, column=col_num, value=h_text)
        cell.font = font_header
        cell.fill = fill_sql_hdr if col_num == 15 else fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # Sample Data for Update Sheet
    sample_updates = [
        [1, "สมุดระบายสีสัตว์น่ารัก เล่ม 1 (ปรับปรุงใหม่)", "", "", "สมุดระบายสีเด็ก", 39, 89, "", "", "", "https://firebasestorage.googleapis.com/.../new_cover1.jpg", "https://firebasestorage.googleapis.com/.../new_book1.pdf", "", ""],
        [2, "ชีทคณิตศาสตร์ คิดเลขเร็ว เล่ม 1", "", "", "คณิตศาสตร์", 49, 119, "", "", "", "", "", "", ""],
        [5, "นิทานเด็ก AI ผจญภัยในอวกาศ", "", "", "นิทานเด็ก AI", 69, 159, "", "", "แนะนำพิเศษ", "https://firebasestorage.googleapis.com/.../cover5.jpg", "", "", ""]
    ]

    for r_idx in range(5, 55): # Rows 5 to 54
        ws_update.row_dimensions[r_idx].height = 22
        if r_idx - 5 < len(sample_updates):
            row_data = sample_updates[r_idx - 5]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_update.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_regular
                cell.border = border_thin
        else:
            for c_idx in range(1, 15):
                ws_update.cell(row=r_idx, column=c_idx).border = border_thin

        # Dynamic Update Formula with _xlfn.TEXTJOIN
        update_formula = (
            f'=IF(ISBLANK(A{r_idx}), "", '
            f'"UPDATE books SET " & _xlfn.TEXTJOIN(", ", TRUE, '
            f'IF(AND(B$3="YES", NOT(ISBLANK(B{r_idx}))), "title=\'" & SUBSTITUTE(B{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(C$3="YES", NOT(ISBLANK(C{r_idx}))), "author_name=\'" & SUBSTITUTE(C{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(D$3="YES", NOT(ISBLANK(D{r_idx}))), "publisher=\'" & SUBSTITUTE(D{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(E$3="YES", NOT(ISBLANK(E{r_idx}))), "category=\'" & SUBSTITUTE(E{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(F$3="YES", NOT(ISBLANK(F{r_idx}))), "price=" & F{r_idx}, ""), '
            f'IF(AND(G$3="YES", NOT(ISBLANK(G{r_idx}))), "original_price=" & G{r_idx}, ""), '
            f'IF(AND(H$3="YES", NOT(ISBLANK(H{r_idx}))), "pages_count=\'" & SUBSTITUTE(H{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(I$3="YES", NOT(ISBLANK(I{r_idx}))), "file_type=\'" & SUBSTITUTE(I{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(J$3="YES", NOT(ISBLANK(J{r_idx}))), "badge=\'" & SUBSTITUTE(J{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(K$3="YES", NOT(ISBLANK(K{r_idx}))), "cover_image_url=\'" & SUBSTITUTE(K{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(L$3="YES", NOT(ISBLANK(L{r_idx}))), "file_path=\'" & SUBSTITUTE(L{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(M$3="YES", NOT(ISBLANK(M{r_idx}))), "sample_file_path=\'" & SUBSTITUTE(M{r_idx}, "\'", "\'\'") & "\'", ""), '
            f'IF(AND(N$3="YES", NOT(ISBLANK(N{r_idx}))), "description=\'" & SUBSTITUTE(N{r_idx}, "\'", "\'\'") & "\'", "")) & '
            f'" WHERE id=" & A{r_idx} & ";")'
        )

        cell_sql = ws_update.cell(row=r_idx, column=15, value=update_formula)
        cell_sql.font = font_code
        cell_sql.fill = fill_sql_bg
        cell_sql.border = border_thin

    # -------------------------------------------------------------
    # 3. DELETE Generator Sheet
    # -------------------------------------------------------------
    ws_delete = wb.create_sheet(title="DELETE_Generator")
    ws_delete.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_delete.merge_cells("A1:I1")
    ws_delete["A1"] = "🗑️ BATCH DELETE SQL GENERATOR (ลบข้อมูลตาม ID หรือเงื่อนไข)"
    ws_delete["A1"].font = font_title
    ws_delete["A1"].fill = fill_title
    ws_delete["A1"].alignment = align_center
    ws_delete.row_dimensions[1].height = 40

    # Section 1: Single ID Delete Table (Cols A-B)
    ws_delete.cell(row=3, column=1, value="📌 วิธีที่ 1: ลบรายหนังสือ (Single Row Delete)").font = font_section
    ws_delete.cell(row=4, column=1, value="book_id (ระบุ ID)").font = font_header
    ws_delete.cell(row=4, column=1).fill = fill_header
    ws_delete.cell(row=4, column=2, value="⚡ GENERATED DELETE STATEMENT").font = font_header
    ws_delete.cell(row=4, column=2).fill = fill_sql_hdr

    sample_delete_ids = [10, 15, 22]
    for r_idx in range(5, 25):
        ws_delete.row_dimensions[r_idx].height = 22
        if r_idx - 5 < len(sample_delete_ids):
            ws_delete.cell(row=r_idx, column=1, value=sample_delete_ids[r_idx - 5]).font = font_regular

        ws_delete.cell(row=r_idx, column=1).border = border_thin
        cell_del_sql = ws_delete.cell(row=r_idx, column=2, value=f'=IF(ISBLANK(A{r_idx}), "", "DELETE FROM books WHERE id = " & A{r_idx} & ";")')
        cell_del_sql.font = font_code
        cell_del_sql.fill = fill_sql_bg
        cell_del_sql.border = border_thin

    # Section 2: Batch IN (...) Delete (Cols D-E)
    ws_delete.cell(row=3, column=4, value="📦 วิธีที่ 2: ลบหลาย ID พร้อมกัน (Batch IN Delete)").font = font_section
    ws_delete.cell(row=4, column=4, value="รายชื่อ ID ที่ต้องการลบ").font = font_header
    ws_delete.cell(row=4, column=4).fill = fill_header

    ws_delete.cell(row=4, column=5, value="⚡ BATCH DELETE IN (...) SQL RESULT").font = font_header
    ws_delete.cell(row=4, column=5).fill = fill_sql_hdr

    batch_ids_sample = [101, 102, 103, 104, 105]
    for r_idx in range(5, 25):
        if r_idx - 5 < len(batch_ids_sample):
            ws_delete.cell(row=r_idx, column=4, value=batch_ids_sample[r_idx - 5]).font = font_regular
        ws_delete.cell(row=r_idx, column=4).border = border_thin

    # Batch IN formula with _xlfn.TEXTJOIN
    batch_in_formula = '="DELETE FROM books WHERE id IN (" & _xlfn.TEXTJOIN(", ", TRUE, D5:D25) & ");"'
    cell_batch_sql = ws_delete.cell(row=5, column=5, value=batch_in_formula)
    cell_batch_sql.font = font_code
    cell_batch_sql.fill = fill_sql_bg
    cell_batch_sql.border = border_thin

    # Section 3: Condition Delete (Cols G-I)
    ws_delete.cell(row=3, column=7, value="🎯 วิธีที่ 3: ลบตามเงื่อนไข (Condition Delete)").font = font_section
    ws_delete.cell(row=4, column=7, value="ฟิลด์เงื่อนไข").font = font_header
    ws_delete.cell(row=4, column=7).fill = fill_header
    ws_delete.cell(row=4, column=8, value="ค่าที่ต้องการระบุ").font = font_header
    ws_delete.cell(row=4, column=8).fill = fill_header
    ws_delete.cell(row=4, column=9, value="⚡ CONDITION DELETE SQL").font = font_header
    ws_delete.cell(row=4, column=9).fill = fill_sql_hdr

    dv_fields = DataValidation(type="list", formula1='"category,publisher,file_type,badge"', allow_blank=False)
    ws_delete.add_data_validation(dv_fields)

    sample_conds = [("category", "สมุดระบายสีเด็ก"), ("badge", "ทดลองใช้")]
    for r_idx in range(5, 15):
        ws_delete.row_dimensions[r_idx].height = 22
        cell_f = ws_delete.cell(row=r_idx, column=7)
        cell_g = ws_delete.cell(row=r_idx, column=8)
        cell_f.border = border_thin
        cell_g.border = border_thin
        dv_fields.add(cell_f)

        if r_idx - 5 < len(sample_conds):
            cell_f.value = sample_conds[r_idx - 5][0]
            cell_g.value = sample_conds[r_idx - 5][1]

        cond_sql = f'=IF(OR(ISBLANK(G{r_idx}), ISBLANK(H{r_idx})), "", "DELETE FROM books WHERE " & G{r_idx} & " = \'" & SUBSTITUTE(H{r_idx}, "\'", "\'\'") & "\';")'
        cell_c_sql = ws_delete.cell(row=r_idx, column=9, value=cond_sql)
        cell_c_sql.font = font_code
        cell_c_sql.fill = fill_sql_bg
        cell_c_sql.border = border_thin

    # Remove default sheet if exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # -------------------------------------------------------------
    # 4. SQL_To_Table_Parser Sheet (Extract DATA VALUES specifically after VALUES clause)
    # -------------------------------------------------------------
    ws_parser = wb.create_sheet(title="SQL_To_Table_Parser")
    ws_parser.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_parser.merge_cells("A1:O1")
    ws_parser["A1"] = "📋 SQL TO TABLE PARSER (แปลงโค้ด SQL ทุกหัวข้อกลับเป็นตารางข้อมูล)"
    ws_parser["A1"].font = font_title
    ws_parser["A1"].fill = fill_title
    ws_parser["A1"].alignment = align_center
    ws_parser.row_dimensions[1].height = 40

    ws_parser["A2"] = "คำแนะนำ: วางโค้ด SQL INSERT หรือโค้ดส่งออกในคอลัมน์ A | ระบบจะสกัดค่าข้อมูลจริง (Data Values) จากชุด VALUES (...) ของทุกหัวข้อกลับเป็นตารางให้อัตโนมัติ"
    ws_parser["A2"].font = font_section

    # Headers for Parser - Covering ALL 14 database data fields!
    headers_parser = [
        "📥 วางโค้ด SQL (Raw SQL Query)", 
        "Parsed ID (คอลัมน์ B)", 
        "Parsed Title (คอลัมน์ C)", 
        "Parsed Author (คอลัมน์ D)", 
        "Parsed Publisher (คอลัมน์ E)", 
        "Parsed Category (คอลัมน์ F)", 
        "Parsed Price (คอลัมน์ G)", 
        "Parsed Original Price (คอลัมน์ H)", 
        "Parsed Pages (คอลัมน์ I)", 
        "Parsed File Type (คอลัมน์ J)", 
        "Parsed Badge (คอลัมน์ K)", 
        "Parsed Cover URL (คอลัมน์ L)", 
        "Parsed File Path (คอลัมน์ M)", 
        "Parsed Sample Path (คอลัมน์ N)", 
        "Parsed Description (คอลัมน์ O)"
    ]

    ws_parser.row_dimensions[4].height = 28
    for col_num, h_text in enumerate(headers_parser, 1):
        cell = ws_parser.cell(row=4, column=col_num, value=h_text)
        cell.font = font_header
        cell.fill = fill_sql_hdr if col_num == 1 else fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # Sample Raw SQL data in Column A
    raw_sqls = [
        "INSERT INTO books (id, title, author_name, publisher, category, price, original_price, pages_count, file_type, badge, cover_image_url, file_path, sample_file_path, description) VALUES (1, 'สมุดระบายสีสัตว์น่ารัก เล่ม 1', 'คลังสมอง', 'คลังสมอง KLANGSAMONG', 'สมุดระบายสีเด็ก', 49, 99, '30 หน้า', 'PDF', 'ขายดี', 'https://firebasestorage.googleapis.com/.../cover1.jpg', 'https://firebasestorage.googleapis.com/.../book1.pdf', 'https://firebasestorage.googleapis.com/.../sample1.pdf', 'แบบฝึกหัดระบายสีสำหรับเด็กปฐมวัย');",
        "INSERT INTO books (id, title, author_name, publisher, category, price, original_price, pages_count, file_type, badge, cover_image_url, file_path, sample_file_path, description) VALUES (2, 'ชีทคณิตศาสตร์ คิดเลขเร็ว เล่ม 1', 'คลังสมอง', 'คลังสมอง KLANGSAMONG', 'คณิตศาสตร์', 59, 129, '45 หน้า', 'PDF', 'แนะนำ', 'https://firebasestorage.googleapis.com/.../cover2.jpg', 'https://firebasestorage.googleapis.com/.../book2.pdf', 'https://firebasestorage.googleapis.com/.../sample2.pdf', 'แบบฝึกหัดคิดเลขเร็ว บวก ลบ เลข 2 หลัก');"
    ]

    for r_idx in range(5, 55):
        ws_parser.row_dimensions[r_idx].height = 22
        if r_idx - 5 < len(raw_sqls):
            ws_parser.cell(row=r_idx, column=1, value=raw_sqls[r_idx - 5]).font = font_code

        ws_parser.cell(row=r_idx, column=1).border = border_thin
        ws_parser.cell(row=r_idx, column=1).fill = fill_sql_bg

        # Expression targeting the VALUES (...) tuple specifically (ignoring column names list)
        v_expr = f'IF(ISNUMBER(SEARCH("VALUES", A{r_idx})), _xlfn.TEXTAFTER(A{r_idx}, "VALUES"), A{r_idx})'

        # Full Formulas for EVERY column from Col B (2) to Col O (15) targeting values tuple:
        f_b = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE({v_expr}, ","), "(", "")))'
        f_c = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 1), ","), "\'", "")))'
        f_d = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 2), ","), "\'", "")))'
        f_e = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 3), ","), "\'", "")))'
        f_f = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 4), ","), "\'", "")))'
        f_g = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 5), ","), "\'", "")))'
        f_h = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 6), ","), "\'", "")))'
        f_i = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 7), ","), "\'", "")))'
        f_j = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 8), ","), "\'", "")))'
        f_k = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 9), ","), "\'", "")))'
        f_l = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 10), ","), "\'", "")))'
        f_m = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 11), ","), "\'", "")))'
        f_n = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER({v_expr}, ",", 12), ","), "\'", "")))'
        f_o = f'=IF(ISBLANK(A{r_idx}), "", TRIM(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(_xlfn.TEXTAFTER({v_expr}, ",", 13), ");", ""), ")", ""), "\'", "")))'

        formulas_all = [f_b, f_c, f_d, f_e, f_f, f_g, f_h, f_i, f_j, f_k, f_l, f_m, f_n, f_o]

        for col_idx, form_text in enumerate(formulas_all, 2):
            cell_parsed = ws_parser.cell(row=r_idx, column=col_idx, value=form_text)
            cell_parsed.font = font_regular
            cell_parsed.border = border_thin

    # -------------------------------------------------------------
    # Adjust Column Widths Across All Worksheets
    # -------------------------------------------------------------
    for ws in [ws_insert, ws_update, ws_delete, ws_parser]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if not val_str.startswith('='):
                    max_len = max(max_len, len(val_str))
            
            if col_letter in ['O', 'N', 'E', 'I', 'L', 'M']:
                ws.column_dimensions[col_letter].width = 45
            elif col_letter in ['A', 'K']:
                ws.column_dimensions[col_letter].width = 30
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # Save workbook safely handling PermissionError if Excel is open
    output_filename = "Batch_SQL_Manager.xlsx"
    try:
        wb.save(output_filename)
        print(f"Successfully updated Excel tool software: {os.path.abspath(output_filename)}")
    except PermissionError:
        output_filename = "Batch_SQL_Manager_Updated.xlsx"
        wb.save(output_filename)
        print(f"Excel file was open. Saved updated Excel tool to: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    create_excel_tool()
