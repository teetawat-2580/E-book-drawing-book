import openpyxl

def verify():
    wb = openpyxl.load_workbook("Batch_SQL_Manager.xlsx", data_only=False)
    print("Sheets in workbook:", wb.sheetnames)
    assert "INSERT_Generator" in wb.sheetnames
    assert "UPDATE_Generator" in wb.sheetnames
    assert "DELETE_Generator" in wb.sheetnames
    assert "SQL_To_Table_Parser" in wb.sheetnames

    # Check _xlfn. prefix on TEXTJOIN / TEXTBEFORE / TEXTAFTER
    ws_ins = wb["INSERT_Generator"]
    formula_ins = ws_ins["A4"].value
    print("INSERT Master Formula A4:", formula_ins)
    assert "_xlfn.TEXTJOIN" in formula_ins

    ws_upd = wb["UPDATE_Generator"]
    formula_upd = ws_upd["O5"].value
    print("UPDATE Formula O5:", formula_upd[:60])
    assert "_xlfn.TEXTJOIN" in formula_upd

    ws_del = wb["DELETE_Generator"]
    formula_del_batch = ws_del["E5"].value
    print("DELETE Batch Formula E5:", formula_del_batch)
    assert "_xlfn.TEXTJOIN" in formula_del_batch

    ws_par = wb["SQL_To_Table_Parser"]
    for col_idx in range(2, 16):
        cell_val = ws_par.cell(row=5, column=col_idx).value
        print(f"Parser Col {col_idx}: {cell_val[:50]}...")
        assert "_xlfn.TEXTBEFORE" in cell_val or "_xlfn.TEXTAFTER" in cell_val

    print("ALL #NAME? FIX VERIFICATIONS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    verify()
